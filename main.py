import openpyxl as xl
from random import choice
from dict import very_bad_malfunction, malfunction
from pathlib import Path
import argparse
import sys

DEFAULT_FILENAME = 'Утиль.xlsx'
DEFAULT_REASON = 'Неустранимые загрязнения'

def get_malfunction_description(model: str) -> str:
    """Возвращает строку с описанием неисправности для указанной модели."""
    try:
        model_malfunctions = malfunction[model]
        return f"{choice(model_malfunctions)}, {choice(very_bad_malfunction)}"
    except KeyError:
        return DEFAULT_REASON

def add_malfunction_reasons(sheet: xl.worksheet.worksheet.Worksheet) -> None:
    """Добавляет причины неисправностей в третий столбец таблицы."""
    for row in range(2, sheet.max_row + 1):
        model = sheet.cell(row=row, column=1).value
        sheet.cell(row=row, column=3).value = get_malfunction_description(model)

def parse_arguments():
    """Парсит аргументы командной строки."""
    parser = argparse.ArgumentParser(
        description='Генератор причин неисправностей для файлов Excel'
    )
    parser.add_argument(
        '-p', '--path',
        default=DEFAULT_FILENAME,
        help=f'Путь к файлу Excel (по умолчанию: {DEFAULT_FILENAME})'
    )
    return parser.parse_args()

def process_workbook(file_path: Path) -> None:
    """Обрабатывает файл Excel."""
    wb = xl.load_workbook(file_path, data_only=True)
    add_malfunction_reasons(wb.active)
    wb.save(file_path)
    wb.close()
    print(f'Файл {file_path} успешно обработан')

def create_default_file(file_path: Path) -> None:
    """Создает файл по умолчанию, если он не существует."""
    wb = xl.Workbook()
    wb.save(file_path)
    wb.close()
    print(f"Создан новый файл: {file_path}")

def main() -> None:
    """Основная функция выполнения скрипта с обработкой исключений."""
    try:
        args = parse_arguments()
        file_path = Path(args.path)
        
        if not file_path.exists():
            create_default_file(file_path)
        
        process_workbook(file_path)
        
    except FileNotFoundError as e:
        print(f'Ошибка: Файл не найден - {e.filename}', file=sys.stderr)
        sys.exit(1)
    except PermissionError as e:
        print(f'Ошибка: Нет прав доступа - {e.filename}', file=sys.stderr)
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print('Ошибка: Неверный формат Excel файла', file=sys.stderr)
        sys.exit(1)
    except KeyError as e:
        print(f'Ошибка в данных: отсутствует ключ - {e}', file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f'Непредвиденная ошибка: {str(e)}', file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
